import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import mysql.connector
import hashlib
from datetime import datetime
from PIL import Image, ImageTk
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# DATABASE INITIALIZATION
def initialize_database():
    try:
        # Connect without database to check if it exists
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password=""
        )
        cur = conn.cursor()
        
        # Create database if not exists
        cur.execute("CREATE DATABASE IF NOT EXISTS expense")
        cur.execute("USE expense")
        
        # Create users table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                fullname VARCHAR(50) NOT NULL,
                username VARCHAR(50) UNIQUE NOT NULL,
                email VARCHAR(100) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                role ENUM('admin','user') DEFAULT 'user'
            )
        """)
        
        # Create categories table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                category_id INT AUTO_INCREMENT PRIMARY KEY,
                category_name VARCHAR(50) UNIQUE NOT NULL
            )
        """)
        
        # Create expenses table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                amount DECIMAL(10,2) NOT NULL,
                spent_on DATE NOT NULL,
                note VARCHAR(255),
                category_ids INT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (category_ids) REFERENCES categories(category_id)
            )
        """)
        
        conn.commit()
        print("Database and tables initialized successfully!")
        
    except Exception as e:
        print(f"Error initializing database: {e}")
    finally:
        conn.close()

# DATABASE CONNECTION
def db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="expense"
    )

# PASSWORD ENCRYPTION
def encrypt_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# INITIALIZE DATABASE WITH CATEGORIES
def initialize_categories():
    try:
        db = db_connection()
        cur = db.cursor()
        categories_list = [
            ('Food',), ('Groceries',), ('Dining Out',), ('Travel',), ('Transport',),
            ('Fuel',), ('Accommodation',), ('Utilities',), ('Electricity',), ('Water',),
            ('Internet',), ('Mobile',), ('Health',), ('Medical',), ('Pharmacy',),
            ('Insurance',), ('Entertainment',), ('Movies',), ('Games',), ('Streaming',),
            ('Shopping',), ('Clothing',), ('Electronics',), ('Education',), ('Books',),
            ('Courses',), ('Fitness',), ('Gym',), ('Hobbies',), ('Gifts',),
            ('Charity',), ('Miscellaneous',)
        ]
        cur.executemany("INSERT IGNORE INTO categories (category_name) VALUES (%s)", categories_list)
        db.commit()
    except Exception as e:
        print(f"Error initializing categories: {e}")
    finally:
        db.close()

# BACKGROUND IMAGE SETUP
bg_label = None

def set_background_image(window):
    global bg_label
    try:
        if bg_label:
            bg_label.destroy()
            
        image_path = r"" # Specify your background image path here currently iam keeping it as empty
        if os.path.exists(image_path):
            window.update_idletasks()
            width = window.winfo_width() if window.winfo_width() >= 400 else 900
            height = window.winfo_height() if window.winfo_height() >= 300 else 650
            
            image = Image.open(image_path)
            image = image.resize((width, height), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(image)
            
            bg_label = tk.Label(window, image=photo)
            bg_label.image = photo
            bg_label.place(x=0, y=0, relwidth=1, relheight=1)
            bg_label.lower()
            return True
        else:
            window.configure(bg='#f0f0f0')
            return False
    except Exception as e:
        window.configure(bg='#f0f0f0')
        return False

def on_window_resize(event):
    if event.widget == window:
        window.after(100, lambda: set_background_image(window))

# NAVIGATION SYSTEM
history = []
forward_stack = []

def navigate_to(page_func, *args):
    global history, forward_stack
    current = window.winfo_children()
    if current:
        history.append(page_func)
    forward_stack.clear()
    for widget in window.winfo_children():
        widget.destroy()
    set_background_image(window)
    page_func(*args)

def go_back():
    global history, forward_stack
    if history:
        current_page = history.pop()
        forward_stack.append(current_page)
        for widget in window.winfo_children():
            widget.destroy()
        set_background_image(window)
        history[-1]() if history else show_main_menu()

# STYLED FRAME
def create_styled_frame(parent, width=400, height=300):
    frame = tk.Frame(parent, bg='white', relief='raised', bd=2)
    frame.configure(highlightbackground='lightgray', highlightthickness=1)
    return frame

# CHECK IF ADMIN EXISTS
def admin_exists():
    try:
        db = db_connection()
        cur = db.cursor()
        cur.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
        count = cur.fetchone()[0]
        return count > 0
    except:
        return False
    finally:
        db.close()

# MAIN MENU
def show_main_menu():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    # Check if admin exists
    has_admin = admin_exists()
    
    main_frame = create_styled_frame(window)
    if has_admin:
        main_frame.place(relx=0.5, rely=0.5, anchor='center', width=350, height=250)
    else:
        main_frame.place(relx=0.5, rely=0.5, anchor='center', width=350, height=290)
    
    tk.Label(main_frame, text="Expense Tracker", font=("Arial", 20, "bold"), bg='white', fg='#2c3e50').pack(pady=20)
    
    if not has_admin:
        tk.Button(main_frame, text="Register as Admin", width=20, command=lambda: navigate_to(register_admin), bg='#e74c3c', fg='white', font=("Arial", 10, "bold")).pack(pady=8)
    
    tk.Button(main_frame, text="Login as Admin", width=20, command=lambda: navigate_to(admin_login), bg='#3498db', fg='white', font=("Arial", 10)).pack(pady=8)
    tk.Button(main_frame, text="Login as User", width=20, command=lambda: navigate_to(user_login), bg='#2ecc71', fg='white', font=("Arial", 10)).pack(pady=8)
    tk.Button(main_frame, text="Register", width=20, command=lambda: navigate_to(register_user), bg='#f39c12', fg='white', font=("Arial", 10)).pack(pady=8)

# REGISTER ADMIN (Only for first-time setup)
def register_admin():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=380)

    tk.Label(main_frame, text="Register First Admin", font=("Arial", 16, "bold"), bg='white', fg='#e74c3c').pack(pady=15)
    tk.Label(main_frame, text="(Only available during first setup)", font=("Arial", 9, "italic"), bg='white', fg='gray').pack()

    tk.Label(main_frame, text="Full Name", bg='white').pack()
    fullname_entry = tk.Entry(main_frame, width=30)
    fullname_entry.pack(pady=5)

    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.pack(pady=5)

    tk.Label(main_frame, text="Email", bg='white').pack()
    email_entry = tk.Entry(main_frame, width=30)
    email_entry.pack(pady=5)
    
    tk.Label(main_frame, text="Password", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def submit():
        # Double-check no admin exists
        if admin_exists():
            messagebox.showerror("Error", "Admin already exists! Only existing admins can create new admins.")
            navigate_to(show_main_menu)
            return
        
        fullname = fullname_entry.get().strip()
        username = username_entry.get().strip()
        email = email_entry.get().strip()
        password = password_entry.get().strip()
        
        if not fullname or not username or not email or not password:
            messagebox.showerror("Error", "All fields are required")
            return

        encrypted = encrypt_password(password)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT email FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                messagebox.showerror("Error", "Email already exists")
                return
            cur.execute("SELECT username FROM users WHERE username=%s", (username,))
            if cur.fetchone():
                messagebox.showerror("Error", "Username already exists")
                return
            
            cur.execute("INSERT INTO users (fullname, username, email, password, role) VALUES (%s, %s, %s, %s, %s)", 
                       (fullname, username, email, encrypted, 'admin'))
            db.commit()
            messagebox.showinfo("Success", "Admin account created successfully! Please login.")
            navigate_to(show_main_menu)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Create Admin Account", command=submit, bg='#e74c3c', fg='white', width=20).pack(pady=15)
    tk.Button(main_frame, text="Back", command=go_back, bg='lightgray', width=20).pack()

# REGISTER USER
def register_user():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=380)

    tk.Label(main_frame, text="Register New User", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=15)

    tk.Label(main_frame, text="Full Name", bg='white').pack()
    fullname_entry = tk.Entry(main_frame, width=30)
    fullname_entry.pack(pady=5)

    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.pack(pady=5)

    tk.Label(main_frame, text="Email", bg='white').pack()
    email_entry = tk.Entry(main_frame, width=30)
    email_entry.pack(pady=5)
    
    tk.Label(main_frame, text="Password", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def submit():
        fullname = fullname_entry.get().strip()
        username = username_entry.get().strip()
        email = email_entry.get().strip()
        password = password_entry.get().strip()
        
        if not fullname or not username or not email or not password:
            messagebox.showerror("Error", "All fields are required")
            return

        encrypted = encrypt_password(password)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT email FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                messagebox.showerror("Error", "Email already exists")
                return
            cur.execute("SELECT username FROM users WHERE username=%s", (username,))
            if cur.fetchone():
                messagebox.showerror("Error", "Username already exists")
                return
            
            cur.execute("INSERT INTO users (fullname, username, email, password, role) VALUES (%s, %s, %s, %s, %s)", 
                       (fullname, username, email, encrypted, 'user'))
            db.commit()
            messagebox.showinfo("Success", "Registration successful! Please login.")
            navigate_to(show_main_menu)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Register", command=submit, bg='#2ecc71', fg='white', width=15).pack(pady=15)
    tk.Button(main_frame, text="Back", command=go_back, bg='lightgray', width=15).pack()

# ADMIN LOGIN
def admin_login():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=300)

    tk.Label(main_frame, text="Admin Login", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=15)
    
    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.pack(pady=5)
    
    tk.Label(main_frame, text="Password", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def login():
        username = username_entry.get().strip()
        password = password_entry.get().strip()
        
        if not username or not password:
            messagebox.showerror("Error", "Please enter both username and password")
            return
            
        encrypted = encrypt_password(password)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT id, role FROM users WHERE username=%s AND password=%s", (username, encrypted))
            result = cur.fetchone()
            if result and result[1] == "admin":
                navigate_to(admin_dashboard, result[0])
            else:
                messagebox.showerror("Error", "Invalid admin credentials")
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Login", command=login, bg='#3498db', fg='white', width=15).pack(pady=15)
    tk.Button(main_frame, text="Back", command=go_back, bg='lightgray', width=15).pack()

# USER LOGIN
def user_login():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=300)

    tk.Label(main_frame, text="User Login", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=15)
    
    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.pack(pady=5)
    
    tk.Label(main_frame, text="Password", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def login():
        username = username_entry.get().strip()
        password = password_entry.get().strip()
        
        if not username or not password:
            messagebox.showerror("Error", "Please enter both username and password")
            return
            
        encrypted = encrypt_password(password)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT id, role, fullname FROM users WHERE username=%s AND password=%s", (username, encrypted))
            result = cur.fetchone()
            if result and result[1] == "user":
                navigate_to(user_dashboard, result[0], result[2])
            else:
                messagebox.showerror("Error", "Invalid user credentials")
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Login", command=login, bg='#2ecc71', fg='white', width=15).pack(pady=15)
    tk.Button(main_frame, text="Back", command=go_back, bg='lightgray', width=15).pack()

# ADMIN DASHBOARD
def admin_dashboard(admin_id):
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)

    container = tk.Frame(window, bg='white', relief='raised', bd=2)
    container.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(container, text="Admin Dashboard", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=10)

    notebook = ttk.Notebook(container)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # Users Management Tab
    users_frame = ttk.Frame(notebook)
    notebook.add(users_frame, text="Users Management")

    # Categories Management Tab
    categories_frame = ttk.Frame(notebook)
    notebook.add(categories_frame, text="Categories")

    # USERS TAB
    tk.Label(users_frame, text="Users Management", font=("Arial", 14, "bold")).pack(pady=5)
    
    users_table = ttk.Treeview(users_frame, columns=("ID", "Fullname", "Username", "Email", "Role"), show="headings", height=12)
    users_table.heading("ID", text="ID")
    users_table.heading("Fullname", text="Full Name")
    users_table.heading("Username", text="Username")
    users_table.heading("Email", text="Email")
    users_table.heading("Role", text="Role")
    users_table.pack(fill="both", expand=True, padx=5, pady=5)

    def refresh_users_table():
        for item in users_table.get_children():
            users_table.delete(item)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT id, fullname, username, email, role FROM users ORDER BY role, fullname")
            for row in cur.fetchall():
                users_table.insert("", "end", values=row)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    refresh_users_table()

    user_buttons_frame = tk.Frame(users_frame)
    user_buttons_frame.pack(pady=10)

    def add_admin():
        navigate_to(admin_create_admin)

    def add_user():
        navigate_to(admin_add_user)

    def edit_user():
        selected = users_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select a user to edit")
            return
        user_data = users_table.item(selected)["values"]
        navigate_to(admin_edit_user, user_data)

    def delete_user():
        selected = users_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select a user")
            return
        user_data = users_table.item(selected)["values"]
        if user_data[4] == 'admin':
            messagebox.showerror("Error", "Cannot delete admin users from here")
            return
        
        if messagebox.askyesno("Confirm", f"Delete user '{user_data[2]}'?"):
            try:
                db = db_connection()
                cur = db.cursor()
                cur.execute("DELETE FROM expenses WHERE user_id=%s", (user_data[0],))
                cur.execute("DELETE FROM users WHERE id=%s", (user_data[0],))
                db.commit()
                messagebox.showinfo("Success", "User deleted")
                refresh_users_table()
            except Exception as e:
                messagebox.showerror("Database Error", str(e))
            finally:
                db.close()

    tk.Button(user_buttons_frame, text="Create New Admin", command=add_admin, width=15, bg='#9b59b6', fg='white').pack(side="left", padx=5)
    tk.Button(user_buttons_frame, text="Add User", command=add_user, width=12, bg='#2ecc71', fg='white').pack(side="left", padx=5)
    tk.Button(user_buttons_frame, text="Edit User", command=edit_user, width=12, bg='#f39c12', fg='white').pack(side="left", padx=5)
    tk.Button(user_buttons_frame, text="Delete User", command=delete_user, width=12, bg='#e74c3c', fg='white').pack(side="left", padx=5)

    # CATEGORIES TAB
    tk.Label(categories_frame, text="Categories Management", font=("Arial", 14, "bold")).pack(pady=5)
    
    categories_table = ttk.Treeview(categories_frame, columns=("ID", "Name"), show="headings", height=12)
    categories_table.heading("ID", text="Category ID")
    categories_table.heading("Name", text="Category Name")
    categories_table.pack(fill="both", expand=True, padx=5, pady=5)

    def refresh_categories_table():
        for item in categories_table.get_children():
            categories_table.delete(item)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT category_id, category_name FROM categories ORDER BY category_name")
            for row in cur.fetchall():
                categories_table.insert("", "end", values=row)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    refresh_categories_table()

    cat_buttons_frame = tk.Frame(categories_frame)
    cat_buttons_frame.pack(pady=10)

    def add_category():
        category_name = tk.simpledialog.askstring("Add Category", "Enter category name:")
        if not category_name:
            return
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT category_id FROM categories WHERE category_name=%s", (category_name.strip(),))
            if cur.fetchone():
                messagebox.showerror("Error", "Category already exists")
                return
            cur.execute("INSERT INTO categories (category_name) VALUES (%s)", (category_name.strip(),))
            db.commit()
            messagebox.showinfo("Success", "Category added")
            refresh_categories_table()
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    def delete_category():
        selected = categories_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select a category")
            return
        cat_data = categories_table.item(selected)["values"]
        
        if messagebox.askyesno("Confirm", f"Delete category '{cat_data[1]}'?"):
            try:
                db = db_connection()
                cur = db.cursor()
                cur.execute("UPDATE expenses SET category_ids=NULL WHERE category_ids=%s", (cat_data[0],))
                cur.execute("DELETE FROM categories WHERE category_id=%s", (cat_data[0],))
                db.commit()
                messagebox.showinfo("Success", "Category deleted")
                refresh_categories_table()
            except Exception as e:
                messagebox.showerror("Database Error", str(e))
            finally:
                db.close()

    tk.Button(cat_buttons_frame, text="Add Category", command=add_category, width=12, bg='#2ecc71', fg='white').pack(side="left", padx=5)
    tk.Button(cat_buttons_frame, text="Delete Category", command=delete_category, width=12, bg='#e74c3c', fg='white').pack(side="left", padx=5)

    nav_frame = tk.Frame(container, bg='white')
    nav_frame.pack(pady=10)
    
    def logout():
        if messagebox.askyesno("Logout", "Are you sure?"):
            navigate_to(show_main_menu)
    
    tk.Button(nav_frame, text="Logout", command=logout, bg='lightgray', width=12).pack()

# ADMIN CREATE NEW ADMIN
def admin_create_admin():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=380)

    tk.Label(main_frame, text="Create New Admin", font=("Arial", 16, "bold"), bg='white', fg='#9b59b6').pack(pady=15)

    tk.Label(main_frame, text="Full Name", bg='white').pack()
    fullname_entry = tk.Entry(main_frame, width=30)
    fullname_entry.pack(pady=5)

    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.pack(pady=5)

    tk.Label(main_frame, text="Email", bg='white').pack()
    email_entry = tk.Entry(main_frame, width=30)
    email_entry.pack(pady=5)
    
    tk.Label(main_frame, text="Password", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def submit():
        fullname = fullname_entry.get().strip()
        username = username_entry.get().strip()
        email = email_entry.get().strip()
        password = password_entry.get().strip()
        
        if not fullname or not username or not email or not password:
            messagebox.showerror("Error", "All fields are required")
            return

        encrypted = encrypt_password(password)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT email FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                messagebox.showerror("Error", "Email already exists")
                return
            cur.execute("SELECT username FROM users WHERE username=%s", (username,))
            if cur.fetchone():
                messagebox.showerror("Error", "Username already exists")
                return
            
            cur.execute("INSERT INTO users (fullname, username, email, password, role) VALUES (%s, %s, %s, %s, %s)", 
                       (fullname, username, email, encrypted, 'admin'))
            db.commit()
            messagebox.showinfo("Success", "New admin account created successfully!")
            navigate_to(admin_dashboard, cur.lastrowid)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Create Admin", command=submit, bg='#9b59b6', fg='white', width=15).pack(pady=15)
    tk.Button(main_frame, text="Cancel", command=lambda: navigate_to(admin_dashboard, 1), bg='lightgray', width=15).pack()

# ADMIN ADD USER
def admin_add_user():
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=380)

    tk.Label(main_frame, text="Add New User", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=15)

    tk.Label(main_frame, text="Full Name", bg='white').pack()
    fullname_entry = tk.Entry(main_frame, width=30)
    fullname_entry.pack(pady=5)

    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.pack(pady=5)

    tk.Label(main_frame, text="Email", bg='white').pack()
    email_entry = tk.Entry(main_frame, width=30)
    email_entry.pack(pady=5)
    
    tk.Label(main_frame, text="Password", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def submit():
        fullname = fullname_entry.get().strip()
        username = username_entry.get().strip()
        email = email_entry.get().strip()
        password = password_entry.get().strip()
        
        if not fullname or not username or not email or not password:
            messagebox.showerror("Error", "All fields are required")
            return

        encrypted = encrypt_password(password)
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("SELECT email FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                messagebox.showerror("Error", "Email already exists")
                return
            cur.execute("SELECT username FROM users WHERE username=%s", (username,))
            if cur.fetchone():
                messagebox.showerror("Error", "Username already exists")
                return
            
            cur.execute("INSERT INTO users (fullname, username, email, password, role) VALUES (%s, %s, %s, %s, %s)", 
                       (fullname, username, email, encrypted, 'user'))
            db.commit()
            messagebox.showinfo("Success", "User added successfully!")
            navigate_to(admin_dashboard, 1)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Add User", command=submit, bg='#2ecc71', fg='white', width=15).pack(pady=15)
    tk.Button(main_frame, text="Cancel", command=lambda: navigate_to(admin_dashboard, 1), bg='lightgray', width=15).pack()

# ADMIN EDIT USER
def admin_edit_user(user_data):
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=400, height=420)

    tk.Label(main_frame, text="Edit User", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=15)

    tk.Label(main_frame, text="Full Name", bg='white').pack()
    fullname_entry = tk.Entry(main_frame, width=30)
    fullname_entry.insert(0, user_data[1])
    fullname_entry.pack(pady=5)

    tk.Label(main_frame, text="Username", bg='white').pack()
    username_entry = tk.Entry(main_frame, width=30)
    username_entry.insert(0, user_data[2])
    username_entry.pack(pady=5)

    tk.Label(main_frame, text="Email", bg='white').pack()
    email_entry = tk.Entry(main_frame, width=30)
    email_entry.insert(0, user_data[3])
    email_entry.pack(pady=5)
    
    tk.Label(main_frame, text="New Password (leave blank to keep current)", bg='white').pack()
    password_entry = tk.Entry(main_frame, width=30, show="*")
    password_entry.pack(pady=5)

    def submit():
        fullname = fullname_entry.get().strip()
        username = username_entry.get().strip()
        email = email_entry.get().strip()
        password = password_entry.get().strip()
        
        if not fullname or not username or not email:
            messagebox.showerror("Error", "Name, username, and email are required")
            return

        try:
            db = db_connection()
            cur = db.cursor()
            
            # Check if email already exists for another user
            cur.execute("SELECT id FROM users WHERE email=%s AND id != %s", (email, user_data[0]))
            if cur.fetchone():
                messagebox.showerror("Error", "Email already exists")
                return
            
            # Check if username already exists for another user
            cur.execute("SELECT id FROM users WHERE username=%s AND id != %s", (username, user_data[0]))
            if cur.fetchone():
                messagebox.showerror("Error", "Username already exists")
                return
            
            # Update user
            if password:
                encrypted = encrypt_password(password)
                cur.execute("UPDATE users SET fullname=%s, username=%s, email=%s, password=%s WHERE id=%s", 
                           (fullname, username, email, encrypted, user_data[0]))
            else:
                cur.execute("UPDATE users SET fullname=%s, username=%s, email=%s WHERE id=%s", 
                           (fullname, username, email, user_data[0]))
            
            db.commit()
            messagebox.showinfo("Success", "User updated successfully!")
            navigate_to(admin_dashboard, 1)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(main_frame, text="Update User", command=submit, bg='#f39c12', fg='white', width=15).pack(pady=15)
    tk.Button(main_frame, text="Cancel", command=lambda: navigate_to(admin_dashboard, 1), bg='lightgray', width=15).pack()

# USER DASHBOARD
def user_dashboard(user_id, fullname):
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    container = tk.Frame(window, bg='white', relief='raised', bd=2)
    container.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(container, text=f"Welcome, {fullname}", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=10)

    notebook = ttk.Notebook(container)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # Add Expense Tab
    add_frame = ttk.Frame(notebook)
    notebook.add(add_frame, text="Add Expense")

    # My Expenses Tab
    view_frame = ttk.Frame(notebook)
    notebook.add(view_frame, text="My Expenses")

    # ADD EXPENSE TAB
    tk.Label(add_frame, text="Add New Expense", font=("Arial", 14, "bold")).pack(pady=15)
    
    form_frame = tk.Frame(add_frame)
    form_frame.pack(pady=10)
    
    tk.Label(form_frame, text="Amount:").grid(row=0, column=0, sticky="e", padx=5, pady=8)
    amount_entry = tk.Entry(form_frame, width=25)
    amount_entry.grid(row=0, column=1, padx=5, pady=8)
    
    tk.Label(form_frame, text="Date (YYYY-MM-DD):").grid(row=1, column=0, sticky="e", padx=5, pady=8)
    date_entry = tk.Entry(form_frame, width=25)
    date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
    date_entry.grid(row=1, column=1, padx=5, pady=8)
    
    tk.Label(form_frame, text="Category:").grid(row=2, column=0, sticky="e", padx=5, pady=8)
    
    try:
        db = db_connection()
        cur = db.cursor()
        cur.execute("SELECT category_id, category_name FROM categories ORDER BY category_name")
        categories = cur.fetchall()
        db.close()
        category_options = [f"{cat[1]}" for cat in categories]
    except:
        categories = []
        category_options = []
    
    category_var = tk.StringVar()
    category_combo = ttk.Combobox(form_frame, textvariable=category_var, values=category_options, width=23)
    category_combo.grid(row=2, column=1, padx=5, pady=8)
    
    tk.Label(form_frame, text="Note:").grid(row=3, column=0, sticky="e", padx=5, pady=8)
    note_entry = tk.Entry(form_frame, width=25)
    note_entry.grid(row=3, column=1, padx=5, pady=8)

    def add_expense():
        try:
            amount = float(amount_entry.get().strip())
            if amount <= 0:
                raise ValueError
        except:
            messagebox.showerror("Error", "Invalid amount")
            return
        
        date_str = date_entry.get().strip()
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except:
            messagebox.showerror("Error", "Invalid date format")
            return
        
        category_name = category_var.get().strip()
        category_id = None
        if category_name:
            for cat in categories:
                if cat[1] == category_name:
                    category_id = cat[0]
                    break
        
        note = note_entry.get().strip()
        
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("INSERT INTO expenses (user_id, amount, spent_on, note, category_ids) VALUES (%s, %s, %s, %s, %s)",
                       (user_id, amount, date_str, note if note else None, category_id))
            db.commit()
            messagebox.showinfo("Success", "Expense added successfully!")
            amount_entry.delete(0, tk.END)
            note_entry.delete(0, tk.END)
            category_var.set('')
            refresh_expenses()
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    tk.Button(form_frame, text="Add Expense", command=add_expense, bg='#2ecc71', fg='white', width=15).grid(row=4, column=0, columnspan=2, pady=20)

    # MY EXPENSES TAB
    tk.Label(view_frame, text="My Expenses", font=("Arial", 14, "bold")).pack(pady=5)
    
    filter_frame = tk.Frame(view_frame)
    filter_frame.pack(pady=5)
    
    tk.Label(filter_frame, text="Filter by Category:").pack(side="left", padx=5)
    filter_var = tk.StringVar(value="All")
    filter_options = ["All"] + category_options
    filter_combo = ttk.Combobox(filter_frame, textvariable=filter_var, values=filter_options, width=15)
    filter_combo.pack(side="left", padx=5)
    
    expenses_table = ttk.Treeview(view_frame, columns=("ID", "Amount", "Date", "Category", "Note"), show="headings", height=10)
    expenses_table.heading("ID", text="ID")
    expenses_table.heading("Amount", text="Amount")
    expenses_table.heading("Date", text="Date")
    expenses_table.heading("Category", text="Category")
    expenses_table.heading("Note", text="Note")
    expenses_table.pack(fill="both", expand=True, padx=5, pady=5)

    total_label = tk.Label(view_frame, text="Total: Rs.0.00", font=("Arial", 12, "bold"), fg='#e74c3c')
    total_label.pack(pady=5)

    def refresh_expenses():
        for item in expenses_table.get_children():
            expenses_table.delete(item)
        
        filter_category = filter_var.get()
        total = 0
        
        try:
            db = db_connection()
            cur = db.cursor()
            
            if filter_category == "All":
                cur.execute("""
                    SELECT e.id, e.amount, e.spent_on, c.category_name, e.note
                    FROM expenses e
                    LEFT JOIN categories c ON e.category_ids = c.category_id
                    WHERE e.user_id = %s
                    ORDER BY e.spent_on DESC
                """, (user_id,))
            else:
                cat_id = None
                for cat in categories:
                    if cat[1] == filter_category:
                        cat_id = cat[0]
                        break
                cur.execute("""
                    SELECT e.id, e.amount, e.spent_on, c.category_name, e.note
                    FROM expenses e
                    LEFT JOIN categories c ON e.category_ids = c.category_id
                    WHERE e.user_id = %s AND e.category_ids = %s
                    ORDER BY e.spent_on DESC
                """, (user_id, cat_id))
            
            for row in cur.fetchall():
                expenses_table.insert("", "end", values=row)
                total += float(row[1])
            
            total_label.config(text=f"Total: Rs.{total:.2f}")
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    def apply_filter():
        refresh_expenses()

    tk.Button(filter_frame, text="Apply Filter", command=apply_filter, bg='#3498db', fg='white').pack(side="left", padx=5)

    refresh_expenses()

    expense_buttons_frame = tk.Frame(view_frame)
    expense_buttons_frame.pack(pady=10)

    def update_expense():
        selected = expenses_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an expense to update")
            return
        
        expense_data = expenses_table.item(selected)["values"]
        navigate_to(user_update_expense, user_id, fullname, expense_data, categories)

    def delete_expense():
        selected = expenses_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an expense")
            return
        
        expense_id = expenses_table.item(selected)["values"][0]
        
        if messagebox.askyesno("Confirm", "Delete this expense?"):
            try:
                db = db_connection()
                cur = db.cursor()
                cur.execute("DELETE FROM expenses WHERE id=%s AND user_id=%s", (expense_id, user_id))
                db.commit()
                messagebox.showinfo("Success", "Expense deleted")
                refresh_expenses()
            except Exception as e:
                messagebox.showerror("Database Error", str(e))
            finally:
                db.close()

    def download_expenses():
        try:
            db = db_connection()
            cur = db.cursor()
            
            # Fetch all expenses for the user
            cur.execute("""
                SELECT e.id, e.amount, e.spent_on, c.category_name, e.note
                FROM expenses e
                LEFT JOIN categories c ON e.category_ids = c.category_id
                WHERE e.user_id = %s
                ORDER BY e.spent_on DESC
            """, (user_id,))
            
            expenses_data = cur.fetchall()
            
            if not expenses_data:
                messagebox.showinfo("No Data", "No expenses to download")
                return
            
            # Ask user where to save
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"expenses_{fullname}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "My Expenses"
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # Add headers
            headers = ["ID", "Amount", "Date", "Category", "Note"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data
            total_amount = 0
            for row_idx, expense in enumerate(expenses_data, start=2):
                ws.cell(row=row_idx, column=1, value=expense[0])
                ws.cell(row=row_idx, column=2, value=float(expense[1]))
                ws.cell(row=row_idx, column=3, value=expense[2].strftime('%Y-%m-%d') if expense[2] else '')
                ws.cell(row=row_idx, column=4, value=expense[3] if expense[3] else 'Uncategorized')
                ws.cell(row=row_idx, column=5, value=expense[4] if expense[4] else '')
                total_amount += float(expense[1])
            
            # Add total row
            total_row = len(expenses_data) + 2
            ws.cell(row=total_row, column=1, value="TOTAL")
            ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=total_row, column=2, value=total_amount)
            ws.cell(row=total_row, column=2).font = Font(bold=True, size=12)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 40
            
            # Save file
            wb.save(file_path)
            messagebox.showinfo("Success", f"Expenses downloaded successfully!\nSaved to: {file_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to download expenses: {str(e)}")
        finally:
            db.close()

    tk.Button(expense_buttons_frame, text="Update Expense", command=update_expense, bg='#f39c12', fg='white', width=15).pack(side="left", padx=5)
    tk.Button(expense_buttons_frame, text="Delete Expense", command=delete_expense, bg='#e74c3c', fg='white', width=15).pack(side="left", padx=5)
    tk.Button(expense_buttons_frame, text="Download to Excel", command=download_expenses, bg='#27ae60', fg='white', width=15).pack(side="left", padx=5)

    nav_frame = tk.Frame(container, bg='white')
    nav_frame.pack(pady=10)
    
    def logout():
        if messagebox.askyesno("Logout", "Are you sure?"):
            navigate_to(show_main_menu)
    
    tk.Button(nav_frame, text="Logout", command=logout, bg='lightgray', width=12).pack()

# USER UPDATE EXPENSE
def user_update_expense(user_id, fullname, expense_data, categories):
    for widget in window.winfo_children():
        widget.destroy()
    
    set_background_image(window)
    
    main_frame = create_styled_frame(window)
    main_frame.place(relx=0.5, rely=0.5, anchor='center', width=450, height=400)

    tk.Label(main_frame, text="Update Expense", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack(pady=15)
    
    form_frame = tk.Frame(main_frame, bg='white')
    form_frame.pack(pady=10)
    
    tk.Label(form_frame, text="Amount:", bg='white').grid(row=0, column=0, sticky="e", padx=5, pady=8)
    amount_entry = tk.Entry(form_frame, width=25)
    amount_entry.insert(0, expense_data[1])
    amount_entry.grid(row=0, column=1, padx=5, pady=8)
    
    tk.Label(form_frame, text="Date (YYYY-MM-DD):", bg='white').grid(row=1, column=0, sticky="e", padx=5, pady=8)
    date_entry = tk.Entry(form_frame, width=25)
    date_entry.insert(0, expense_data[2])
    date_entry.grid(row=1, column=1, padx=5, pady=8)
    
    tk.Label(form_frame, text="Category:", bg='white').grid(row=2, column=0, sticky="e", padx=5, pady=8)
    
    category_options = [f"{cat[1]}" for cat in categories]
    category_var = tk.StringVar(value=expense_data[3] if expense_data[3] else "")
    category_combo = ttk.Combobox(form_frame, textvariable=category_var, values=category_options, width=23)
    category_combo.grid(row=2, column=1, padx=5, pady=8)
    
    tk.Label(form_frame, text="Note:", bg='white').grid(row=3, column=0, sticky="e", padx=5, pady=8)
    note_entry = tk.Entry(form_frame, width=25)
    note_entry.insert(0, expense_data[4] if expense_data[4] else "")
    note_entry.grid(row=3, column=1, padx=5, pady=8)

    def submit():
        try:
            amount = float(amount_entry.get().strip())
            if amount <= 0:
                raise ValueError
        except:
            messagebox.showerror("Error", "Invalid amount")
            return
        
        date_str = date_entry.get().strip()
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except:
            messagebox.showerror("Error", "Invalid date format")
            return
        
        category_name = category_var.get().strip()
        category_id = None
        if category_name:
            for cat in categories:
                if cat[1] == category_name:
                    category_id = cat[0]
                    break
        
        note = note_entry.get().strip()
        
        try:
            db = db_connection()
            cur = db.cursor()
            cur.execute("UPDATE expenses SET amount=%s, spent_on=%s, note=%s, category_ids=%s WHERE id=%s AND user_id=%s",
                       (amount, date_str, note if note else None, category_id, expense_data[0], user_id))
            db.commit()
            messagebox.showinfo("Success", "Expense updated successfully!")
            navigate_to(user_dashboard, user_id, fullname)
        except Exception as e:
            messagebox.showerror("Database Error", str(e))
        finally:
            db.close()

    button_frame = tk.Frame(main_frame, bg='white')
    button_frame.pack(pady=20)
    tk.Button(button_frame, text="Update", command=submit, bg='#f39c12', fg='white', width=15).pack(side="left", padx=5)
    tk.Button(button_frame, text="Cancel", command=lambda: navigate_to(user_dashboard, user_id, fullname), bg='lightgray', width=15).pack(side="left", padx=5)

# MAIN APPLICATION
if __name__ == "__main__":
    import tkinter.simpledialog
    
    # Initialize database and tables on startup
    initialize_database()
    
    # Initialize categories on startup
    initialize_categories()
    
    window = tk.Tk()
    window.title("Expense Tracker System")
    window.geometry("950x680")
    window.minsize(800, 600)
    
    window.bind("<Configure>", on_window_resize)
    
    show_main_menu()
    
    window.mainloop()